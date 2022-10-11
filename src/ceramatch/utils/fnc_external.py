from openpyxl import (load_workbook, Workbook)
from openpyxl.styles import (Font)
from collections import defaultdict
import csv
import os

def save_xlsx(header, rows, path):
	# header = [column, ...]
	# rows = [[value, ...], ...]
	
	wb = Workbook()
	wb.guess_types = True
	ws = wb.active
	ws.append(header)
	for i in range(len(header)):
		ws.cell(row = 1, column = i + 1).font = Font(bold = True)
	r = 2
	for row in rows:
		ws.append(row)
		for i in range(len(row)):
			ws.cell(row = r, column = i + 1).number_format = "@"
		r += 1
	wb.save(path)

def save_csv(header, rows, path):
	# header = [column, ...]
	# rows = [[value, ...], ...]
	
	with open(path, "w", newline = "") as f:
		writer = csv.writer(f, dialect=csv.excel, quoting=csv.QUOTE_ALL)
		writer.writerow(header)
		for row in rows:
			writer.writerow([str(val) for val in row])

def import_clusters_xlsx(path, sample_column, cluster_column):
	
	wb = load_workbook(filename = path, read_only = True)
	ws = None
	for sheet in wb.sheetnames:
		ws = wb[sheet]
		break
	if ws is None:
		return {}
	columns = {}
	for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
		value = cell.value
		if value is not None:
			value = str(value).strip()
			if value:
				columns[value] = i
	if (sample_column not in columns) or (cluster_column not in columns):
		return {}
	
	clusters = defaultdict(set)
	for row in ws.iter_rows(min_row = 2):
		sample_id = str(row[columns[sample_column]].value).strip()
		cluster = str(row[columns[cluster_column]].value).strip()
		if sample_id and cluster:
			clusters[cluster].add(sample_id)
	
	return dict(clusters)

def import_clusters_csv(path, sample_column, cluster_column):
	
	columns = {}
	with open(path, "r", newline = "") as f:
		reader = csv.reader(f, dialect = csv.excel, quoting=csv.QUOTE_ALL)
		for row in reader:
			for i, value in enumerate(row):
				value = value.strip()
				if value:
					columns[value] = i
			break
	if (sample_column not in columns) or (cluster_column not in columns):
		return {}
	
	clusters = defaultdict(set)
	with open(path, "r", newline = "") as f:
		reader = csv.reader(f, dialect = csv.excel, quoting=csv.QUOTE_ALL)
		next(reader)
		for row in reader:
			sample_id = str(row[columns[sample_column]]).strip()
			cluster = str(row[columns[cluster_column]]).strip()
			if sample_id and cluster:
				clusters[cluster].add(sample_id)
	
	return dict(clusters)

