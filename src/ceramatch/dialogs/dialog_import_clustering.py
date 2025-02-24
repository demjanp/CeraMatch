from PySide6 import (QtWidgets, QtCore, QtGui)
from openpyxl import load_workbook
from pathlib import Path
import csv
import os

def get_columns_xlsx(path):
	
	wb = load_workbook(filename = path, read_only = True)
	columns = []
	for sheet in wb.sheetnames:
		ws = wb[sheet]
		for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
			value = cell.value
			if value is not None:
				value = str(value).strip()
				if value:
					columns.append(value)
		break
	return columns

def get_columns_csv(path):
	
	columns = []
	with open(path, "r", newline = "") as f:
		reader = csv.reader(f, dialect = csv.excel, quoting=csv.QUOTE_ALL)
		for row in reader:
			for i, value in enumerate(row):
				value = value.strip()
				if value:
					columns.append(value)
			break
	return columns

class DialogImportClustering(QtWidgets.QFrame):
	
	def __init__(self, cview):
		
		QtWidgets.QFrame.__init__(self)
		
		self.cview = cview
		
		self.setMinimumWidth(512)
		self.setLayout(QtWidgets.QFormLayout())
		
		self.path_edit = QtWidgets.QLineEdit()
		self.path_edit.textChanged.connect(self.on_path_changed)
		path_button = QtWidgets.QPushButton("Browse")
		path_button.setSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
		path_button.clicked.connect(self.on_path_button)
		path_frame = QtWidgets.QWidget()
		path_frame.setLayout(QtWidgets.QHBoxLayout())
		path_frame.layout().setContentsMargins(0, 0, 0, 0)
		path_frame.layout().addWidget(self.path_edit)
		path_frame.layout().addWidget(path_button)
		
		self.sample_combo = QtWidgets.QComboBox()
		self.cluster_combo = QtWidgets.QComboBox()
		
		self.layout().addRow("Source File:", path_frame)
		self.layout().addRow("Sample Column:", self.sample_combo)
		self.layout().addRow("Cluster Column:", self.cluster_combo)
	
	@QtCore.Slot()
	def on_path_changed(self):
		
		path = self.path_edit.text()
		if not path:
			return
		if not os.path.isfile(path):
			return
		ext = os.path.splitext(path)[-1].lower()
		columns = []
		if ext == ".xlsx":
			columns = get_columns_xlsx(path)
		elif ext == ".csv":
			columns = get_columns_csv(path)
		if columns:
			self.sample_combo.clear()
			self.sample_combo.addItems(columns)
			self.cluster_combo.clear()
			self.cluster_combo.addItems(columns)
			if len(columns) > 1:
				self.cluster_combo.setCurrentIndex(1)
	
	@QtCore.Slot()
	def on_path_button(self):
		
		path, format = self.cview.get_load_path(
			"Select Source File",
			"Excel 2007+ Workbook (*.xlsx);;Comma-separated Values (*.csv)",
		)
		if path is None:
			return
		self.path_edit.setText(path)

