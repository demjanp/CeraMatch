
from lib.fnc_matching import *

from deposit.DModule import (DModule)
from deposit.store.DLabel.DGeometry import (DGeometry)

from openpyxl import (load_workbook, Workbook)
from openpyxl.styles import (Font)
from collections import defaultdict
from natsort import natsorted
import csv
import os

class Clusters(DModule):
	
	def __init__(self, model):
		
		self.model = model
		
		self._node_obj_lookup = {}  # {node_id: obj_id, ...}
		self._has_clusters = False
		
		DModule.__init__(self)
	
	def has_clusters(self):
		
		return self._has_clusters
	
	def clear(self):
		
		self._has_clusters = False
		self._node_obj_lookup = dict([(node_id, self._node_obj_lookup[node_id]) for node_id in self._node_obj_lookup if not node_id.startswith("#")])
	
	def update_samples(self):
		
		if not self.model.sample_data:
			self._node_obj_lookup = {}
		else:
			self._node_obj_lookup = dict([("@%s" % (sample_id), self.model.sample_data[sample_id][0]) for sample_id in self.model.sample_data])
	
	def format_node_id(self, idx):
		
		if idx < len(self.model.sample_ids):
			return "@%s" % (self.model.sample_ids[idx])
		else:
			return "#%d" % (idx)
	
	def format_node_ids(self, clusters, nodes, edges, labels, positions = None):
		
		clusters = dict([(self.format_node_id(idx), [self.format_node_id(i) for i in clusters[idx]]) for idx in clusters])
		nodes = set([self.format_node_id(idx) for idx in nodes])
		edges = set([(self.format_node_id(idx1), self.format_node_id(idx2)) for idx1, idx2 in edges])
		labels = dict([(self.format_node_id(idx), labels[idx]) for idx in labels])
		if positions is None:
			return clusters, nodes, edges, labels
		else:
			positions = dict([(self.format_node_id(idx), positions[idx]) for idx in positions])
			return clusters, nodes, edges, labels, positions
	
	def load(self):
		# returns clusters, nodes, edges, labels, positions
		#
		# clusters = {#node_id: [@sample_id, ...], ...}
		# nodes = [@sample_id, #node_id, ...]
		# edges = [(@sample_id, #node_id), (#node_id1, #node_id2), ...]
		# labels = {@sample_id: label, #node_id: label, ...}
		# positions = {@sample_id: (x, y), #node_id: (x, y), ...}
		
		def _get_position(obj, node_idx, positions):
			
			try:
				pos = obj.descriptors["CMPosition"].label.coords[0][0]
			except:
				pos = None
			if pos is not None:
				positions[node_idx] = pos
		
		self.clear()
		
		if (not self.model.has_samples()) or (not self.model.has_cluster_classes()):
			return None, None, None, None, None
		
		obj_lookup = {}
		sample_cls, id_descr = self.model.lap_descriptors["Custom_Id"]
		for obj_id in self.model.classes[sample_cls].objects:
			sample_id = self.model.objects[obj_id].descriptors[id_descr].label.value
			if sample_id not in self.model.sample_ids:
				continue
			obj_lookup[obj_id] = self.model.sample_ids.index(sample_id)
		
		positions = {}
		
		nodes = set(list(range(len(self.model.sample_ids))))
		for sample_idx in nodes:
			obj_sample = self.model.objects[self.model.sample_data[self.model.sample_ids[sample_idx]][0]]
			_get_position(obj_sample, sample_idx, positions)
		
		edges = set([])
		clusters = {}
		labels = {}
		node_idx = max(nodes)
		node_idx_lookup = {}  # {obj_id: node_idx, ...}
		for obj_id in self.model.classes[self.model.cluster_class].objects:
			obj_clu = self.model.objects[obj_id]
			if "contains" not in obj_clu.relations:
				continue
			node_idx += 1
			node_idx_lookup[obj_id] = node_idx
			clusters[node_idx] = []
			labels[node_idx] = obj_clu.descriptors["Name"].label.value
			_get_position(obj_clu, node_idx, positions)
			nodes.add(node_idx)
			for obj_id_sample in obj_clu.relations["contains"]:
				sample_idx = obj_lookup[obj_id_sample]
				node_idx_lookup[obj_id_sample] = sample_idx
				clusters[node_idx].append(sample_idx)
				labels[sample_idx] = self.model.sample_ids[sample_idx]
				_get_position(self.model.objects[obj_id_sample], sample_idx, positions)
				nodes.add(sample_idx)
				edges.add((node_idx, sample_idx))
		for obj_id in self.model.classes[self.model.node_class].objects:
			obj_node = self.model.objects[obj_id]
			if self.model.cluster_class in obj_node.classes:
				continue
			node_idx += 1
			node_idx_lookup[obj_id] = node_idx
			labels[node_idx] = obj_node.descriptors["Name"].label.value
			_get_position(obj_node, node_idx, positions)
			nodes.add(node_idx)
		for obj_id1 in self.model.classes[self.model.node_class].objects:
			for obj_id2 in self.model.objects[obj_id1].relations["linked"]:
				edges.add((node_idx_lookup[obj_id1], node_idx_lookup[obj_id2]))
		if clusters:
			self._has_clusters = True
			for obj_id in node_idx_lookup:
				self._node_obj_lookup[self.format_node_id(node_idx_lookup[obj_id])] = obj_id
		
		return self.format_node_ids(clusters, nodes, edges, labels, positions)
	
	def store(self, clusters, nodes, edges, labels):
		
		if not self.model.has_cluster_classes():
			return
		
		node_class = self.model.classes.add(self.model.node_class)
		cluster_class = self.model.classes.add(self.model.cluster_class)
		for node_id in clusters:
			obj_clu = cluster_class.objects.add()
			obj_clu.add_class(node_class)
			self._node_obj_lookup[node_id] = obj_clu.id
			obj_clu.add_descriptor("Name", labels[node_id])
			for sample_node_id in clusters[node_id]:
				obj_clu.add_relation("contains", self.model.sample_data[sample_node_id[1:]][0])
		for node_id in nodes:
			if node_id.startswith("@"):
				continue
			if node_id in self._node_obj_lookup:
				continue
			obj_node = node_class.objects.add()
			self._node_obj_lookup[node_id] = obj_node.id
			obj_node.add_descriptor("Name", labels[node_id])
		for node_id1, node_id2 in edges:
			if (node_id1 in self._node_obj_lookup) and (node_id2 in self._node_obj_lookup):
				self.model.objects[self._node_obj_lookup[node_id1]].add_relation("linked", self._node_obj_lookup[node_id2])
	
	def store_position(self, node_id, x, y):
		
		if node_id not in self._node_obj_lookup:
			return
		obj = self.model.objects[self._node_obj_lookup[node_id]]
		obj.add_descriptor("CMPosition", DGeometry([[x,y]]))
	
	def delete(self):
		
		self.clear()
		if not self.model.has_cluster_classes():
			return
		for cls in [self.model.cluster_class, self.model.node_class]:
			del self.model.classes[cls]
	
	def make(self, max_clusters, limit):
		# returns clusters, nodes, edges, labels
		#
		# clusters = {#node_id: [@sample_id, ...], ...}
		# nodes = [@sample_id, #node_id, ...]
		# edges = [(@sample_id, #node_id), (#node_id1, #node_id2), ...]
		# labels = {@sample_id: label, #node_id: label, ...}
		
		if not self.model.has_cluster_classes():
			return None, None, None, None
		
		self.delete()
		
		if not self.model.has_distance():
			return None, None, None, None
		
		clusters, nodes, edges, labels = get_clusters(self.model.distance, max_clusters = max_clusters, limit = limit)
		if not clusters:
			return None, None, None, None
		
		self._has_clusters = True
		clusters, nodes, edges, labels = self.format_node_ids(clusters, nodes, edges, labels)
		self.store(clusters, nodes, edges, labels)
		
		return clusters, nodes, edges, labels
	
	def update(self):
		# returns clusters, nodes, edges, labels
		#
		# clusters = {#node_id: [@sample_id, ...], ...}
		# nodes = [@sample_id, #node_id, ...]
		# edges = [(@sample_id, #node_id), (#node_id1, #node_id2), ...]
		# labels = {@sample_id: label, #node_id: label, ...}
		
		if not self.model.has_cluster_classes():
			return None, None, None, None
		
		if not self.model.has_distance():
			return None, None, None, None
		
		clusters, _, _, _, _ = self.load()
		
		self.delete()
		
		if clusters:
			# update clusters
			clusters = dict([(label, [self.model.sample_ids.index(sample_node_id[1:]) for sample_node_id in clusters[label]]) for label in clusters])
			clusters, nodes, edges, labels = update_clusters(self.model.distance, clusters)
			if not clusters:
				return None, None, None, None
			self._has_clusters = True
			clusters, nodes, edges, labels = self.format_node_ids(clusters, nodes, edges, labels)
			self.store(clusters, nodes, edges, labels)
			
		else:
			# display tree without clusters
			self._has_clusters = False
			clusters, nodes, edges, labels = get_clusters(self.model.distance, max_clusters = len(self.model.sample_ids))
			nodes_clu = dict([(cluster_id, [-1, -1]) for cluster_id in clusters]) # {cluster_id: [node_id, node_out], ...}
			for node_id1, node_id2 in edges:
				if node_id1 in clusters:
					nodes_clu[node_id1][1] = node_id2
				elif node_id2 in clusters:
					nodes_clu[node_id2][0] = node_id1
			for cluster_id in nodes_clu:
				nodes.remove(cluster_id)
				del labels[cluster_id]
				node_in, node_out = nodes_clu[cluster_id]
				edges.remove((node_in, cluster_id))
				edges.remove((cluster_id, node_out))
				edges.add((node_in, node_out))
			clusters = {}
			clusters, nodes, edges, labels = self.format_node_ids(clusters, nodes, edges, labels)
		
		return clusters, nodes, edges, labels
	
	def add_cluster(self, cluster_id, node_ids):
		
		if not self.model.has_cluster_classes():
			return
		node_class = self.model.classes.add(self.model.node_class)
		cluster_class = self.model.classes.add(self.model.cluster_class)
		
		obj_clu = cluster_class.objects.add()
		obj_clu.add_class(node_class)
		self._node_obj_lookup[cluster_id] = obj_clu.id
		obj_clu.add_descriptor("Name", "Manual")
		for sample_node_id in node_ids:
			obj_clu.add_relation("contains", self.model.sample_data[sample_node_id[1:]][0])
			obj_clu.add_relation("linked", self.model.sample_data[sample_node_id[1:]][0])
	
	def remove(self, node_id):
		
		if node_id not in self._node_obj_lookup:
			return
		del self.model.objects[self._node_obj_lookup[node_id]]
		del self._node_obj_lookup[node_id]
	
	def add_child(self, cluster_id, node_id):
		
		if (cluster_id not in self._node_obj_lookup) or (node_id not in self._node_obj_lookup):
			return
		self.model.objects[self._node_obj_lookup[cluster_id]].add_relation("contains", self._node_obj_lookup[node_id])
	
	def remove_child(self, cluster_id, node_id):
		
		if (cluster_id not in self._node_obj_lookup) or (node_id not in self._node_obj_lookup):
			return
		self.model.objects[self._node_obj_lookup[cluster_id]].del_relation("contains", self._node_obj_lookup[node_id])
	
	def get_cluster_data(self):
		# returns [[sample_id, cluster_name], ...]
		
		sample_class, id_descr = self.model.lap_descriptors["Custom_Id"]
		
		if not self.model.cluster_class:
			return []
		
		data = []
		for clu_id in self.model.classes[self.model.cluster_class].objects:
			obj_clu = self.model.objects[clu_id]
			name = obj_clu.descriptors["Name"].label.value
			for obj_id in obj_clu.relations["contains"]:
				obj_sample = self.model.objects[obj_id]
				if not sample_class in obj_sample.classes:
					continue
				sample_id = obj_sample.descriptors[id_descr].label.value
				data.append([sample_id, name])
		data = natsorted(data)
		return data
	
	def set_cluster_data(self, data):
		# data = [(sample_id, label), ...]
		# returns clusters, nodes, edges, labels
		
		self.delete()
		
		if not self.model.has_cluster_classes():
			return None, None, None, None
		
		if not self.model.has_distance():
			return None, None, None, None
		
		clusters = defaultdict(set)    # {label: [sample_id, ...], ...}
		label_lookup = {}
		for sample_id, label in data:
			if sample_id in self.model.sample_ids:
				sample_idx = self.model.sample_ids.index(sample_id)
				clusters[label].add(sample_idx)
				label_lookup[sample_idx] = label
		clusters = dict([(idx, list(clusters[label])) for idx, label in enumerate(clusters.keys())])
		clu_idx = max(list(clusters.keys())) + 1
		for sample_idx in range(len(self.model.sample_ids)):
			if sample_idx not in label_lookup:
				clusters[clu_idx] = [sample_idx]
				clu_idx += 1
		clusters, nodes, edges, labels = update_clusters(self.model.distance, clusters)
		if not clusters:
			return None, None, None, None
		for idx in clusters:
			sample_idx = clusters[idx][0]
			if sample_idx in label_lookup:
				labels[idx] = label_lookup[sample_idx]
		
		self._has_clusters = True
		clusters, nodes, edges, labels = self.format_node_ids(clusters, nodes, edges, labels)
		self.store(clusters, nodes, edges, labels)
		
		return clusters, nodes, edges, labels
	
	def export_xlsx(self, path):
		
		data = self.get_cluster_data()
		if not data:
			return
		
		wb = Workbook()
		wb.guess_types = True
		ws = wb.active
		ws.append(["Sample ID", "Cluster"])
		for i in range(2):
			ws.cell(row = 1, column = i + 1).font = Font(bold = True)
		row = 2
		for sample_id, cluster in data:
			ws.append([sample_id, cluster])
			for i in range(2):
				ws.cell(row = row, column = i + 1).number_format = "@"
			row += 1
		wb.save(path)
	
	def export_csv(self, path):
		
		data = self.get_cluster_data()
		if not data:
			return
		
		with open(path, "w", newline = "") as f:
			writer = csv.writer(f, dialect=csv.excel, quoting=csv.QUOTE_ALL)
			writer.writerow(["Sample ID", "Cluster"])
			for sample_id, cluster in data:
				writer.writerow([str(sample_id), str(cluster)])
	
	def import_xlsx(self, path, sample_column, cluster_column):
		
		if not os.path.isfile(path):
			return
		wb = load_workbook(filename = path, read_only = True)
		ws = None
		for sheet in wb.sheetnames:
			ws = wb[sheet]
			break
		if ws is None:
			return
		columns = {}
		for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
			value = cell.value
			if value is not None:
				value = str(value).strip()
				if value:
					columns[value] = i
		if (sample_column not in columns) or (cluster_column not in columns):
			return
		data = set([])
		for row in ws.iter_rows(min_row = 2):
			sample_id = str(row[columns[sample_column]].value).strip()
			cluster = str(row[columns[cluster_column]].value).strip()
			if sample_id and cluster:
				data.add((sample_id, cluster))
		
		return self.set_cluster_data(data)
	
	def import_csv(self, path, sample_column, cluster_column):
		
		if not os.path.isfile(path):
			return
		columns = {}
		with open(path, "r", newline = "") as f:
			reader = csv.reader(f, dialect = csv.excel, quoting=csv.QUOTE_ALL)
			for row in reader:
				for i, value in enumerate(row):
					value = value.strip()
					if value:
						columns[value] = i
				break
		if (sample_column not in columns) or (cluster_column not in columns):
			return
		data = set([])
		with open(path, "r", newline = "") as f:
			reader = csv.reader(f, dialect = csv.excel, quoting=csv.QUOTE_ALL)
			for row in reader:
				sample_id = str(row[columns[sample_column]]).strip()
				cluster = str(row[columns[cluster_column]]).strip()
				if sample_id and cluster:
					data.add((sample_id, cluster))
		
		return self.set_cluster_data(data)
	