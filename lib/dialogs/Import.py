
from lib.dialogs._Dialog import (Dialog)

from PySide2 import (QtWidgets, QtCore, QtGui)
from openpyxl import load_workbook
from pathlib import Path
import csv
import os

def get_columns_xlsx(path):
	
	wb = load_workbook(filename = path, read_only = True)
	columns = []
	for sheet in wb.sheetnames:
		ws = wb[sheet]
		for i, cell in enumerate(list(ws.iter_rows(max_row = 1))[0]):
			value = cell.value
			if value is not None:
				value = str(value).strip()
				if value:
					columns.append(value)
		break
	return columns

def get_columns_csv(path):
	
	columns = []
	with open(path, "r", newline = "") as f:
		reader = csv.reader(f, dialect = csv.excel, quoting=csv.QUOTE_ALL)
		for row in reader:
			for i, value in enumerate(row):
				value = value.strip()
				if value:
					columns.append(value)
			break
	return columns

class Import(Dialog):
	
	def title(self):
		
		return "Import Clustering"
	
	def set_up(self):
		
		
		self.setMinimumWidth(512)
		self.setModal(True)
		self.setLayout(QtWidgets.QFormLayout())
		
		self.path_edit = QtWidgets.QLineEdit()
		self.path_edit.textChanged.connect(self.on_path_changed)
		path_button = QtWidgets.QPushButton("Browse")
		path_button.setSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
		path_button.clicked.connect(self.on_path_button)
		path_frame = QtWidgets.QWidget()
		path_frame.setLayout(QtWidgets.QHBoxLayout())
		path_frame.layout().setContentsMargins(0, 0, 0, 0)
		path_frame.layout().addWidget(self.path_edit)
		path_frame.layout().addWidget(path_button)
		
		self.sample_combo = QtWidgets.QComboBox()
		self.cluster_combo = QtWidgets.QComboBox()
		
		self.layout().addRow("Source File:", path_frame)
		self.layout().addRow("Sample Column:", self.sample_combo)
		self.layout().addRow("Cluster Column:", self.cluster_combo)
	
	def process(self):
		
		path = self.path_edit.text().strip()
		sample_column = self.sample_combo.currentText()
		cluster_column = self.cluster_combo.currentText()
		if (not sample_column) or (not cluster_column) or (not os.path.isfile(path)):
			return
		ext = os.path.splitext(path)[-1].lower()
		if ext == ".xlsx":
			clusters, nodes, edges, labels = self.model.clusters.import_xlsx(path, sample_column, cluster_column)
		elif ext == ".csv":
			clusters, nodes, edges, labels = self.model.clusters.import_csv(path, sample_column, cluster_column)
		else:
			return
		self.view.set_clusters(clusters, nodes, edges, labels)
	
	@QtCore.Slot()
	def on_path_changed(self):
		
		path = self.path_edit.text()
		if not path:
			return
		if not os.path.isfile(path):
			return
		ext = os.path.splitext(path)[-1].lower()
		columns = []
		if ext == ".xlsx":
			columns = get_columns_xlsx(path)
		elif ext == ".csv":
			columns = get_columns_csv(path)
		if columns:
			self.sample_combo.clear()
			self.sample_combo.addItems(columns)
			self.cluster_combo.clear()
			self.cluster_combo.addItems(columns)
			if len(columns) > 1:
				self.cluster_combo.setCurrentIndex(1)
	
	@QtCore.Slot()
	def on_path_button(self):
		
		format_xlsx = "Excel 2007+ Workbook (*.xlsx)"
		format_csv = "Comma-separated Values (*.csv)"
		formats = ";;".join([format_xlsx, format_csv])
		default_path = os.path.join(str(Path.home()), "Desktop")
		path, _ = QtWidgets.QFileDialog.getOpenFileName(None, caption = "Select Source File", dir = default_path, filter = formats)
		if not path:
			return
		self.path_edit.setText(path)

